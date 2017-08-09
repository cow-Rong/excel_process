# coding:utf8
import sys
reload(sys)
sys.setdefaultencoding('utf-8')
from openpyxl import load_workbook

wb = load_workbook('f:/pic_keyword.xlsx')
ws = wb.active

f = open('f:/test.txt','w')
for i in range(2,4445):
    a = ws['D'+str(i)].value
    pi_id = str(ws['A'+str(i)].value)
    cn_title = '\"'+str(ws['D'+str(i)].value)+'\"'
    cn_keywords = '\"'+str(ws['E'+str(i)].value)+'\"'
    en_title = '\"'+str(ws['F'+str(i)].value)+'\"'
    en_keywords = '\"'+str(ws['G'+str(i)].value)+'\"'
    str1 = 'UPDATE pic_priminfo SET PPI_Name='+cn_title+', PPI_Intro='+cn_keywords+' WHERE PI_ID='+pi_id+';\n'
    str2 = 'UPDATE pic_priminfo_en SET PPI_Name='+en_title+', PPI_Intro='+en_keywords+' WHERE PI_ID='+pi_id+';\n'
    f.write(str1)
    f.write(str2)
f.close()