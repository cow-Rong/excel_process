# coding:utf8

from openpyxl import Workbook
import pymysql.cursors

# 查询需要修改关键字的图片
def select_pics(connection,user_id):
    try:
        with connection.cursor() as cursor:
            sql = """SELECT a.PI_ID, a.PI_Code, LEFT(a.E_Time, 10) AS E_Time, b.PPI_Name AS CN_Name, b.PPI_Intro AS CN_Intro, c.PPI_Name AS EN_Name, c.PPI_Intro AS EN_Intro,  
            a.PI_State, d.UI_ID, e.UA_SerName, e.UA_Name 
            FROM pic_index a 
            LEFT JOIN pic_priminfo b ON a.PI_ID=b.PI_ID 
            LEFT JOIN pic_priminfo_en c ON a.PI_ID=c.PI_ID
            INNER JOIN user_index d ON a.E_User=d.UI_ID   
            INNER JOIN user_authinfo e ON e.UI_ID=d.UI_ID   
            WHERE b.PPI_State=1 AND a.E_User=4225 AND a.PI_State=1 
            ORDER BY PI_ID"""
            cursor.execute(sql,(user_id))
            return cursor.fetchall()
    except Exception:
        print 'mysql error'

connection = pymysql.connect(host='localhost', user='root', password='123456', db='origino', charset='utf8',cursorclass=pymysql.cursors.DictCursor)
pics = select_pics(connection, user_id='123')
connection.close()
wb = Workbook()
ws = wb.active
for pic in pics:
    pic['asd'] = 3
ws['A1'] = 1
ws['A2'] = 2
ws['A3'] = 3
ws['A4'] = 4

wb.save('f:/balances.xlsx')