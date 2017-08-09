# coding:utf8
import sys
reload(sys)
sys.setdefaultencoding('utf-8')
from openpyxl import load_workbook
import pymysql.cursors

# 查询图片详情
def select_pics(connection,storage):
    try:
        with connection.cursor() as cursor:
            sql = """SELECT e.PI_ID,f.PPI_Intro AS CN_Intro,f.PPI_Name AS CN_Name, g.PPI_Intro AS EN_Intro, g.PPI_Name AS EN_Name, h.UR_Name, i.PT_Name AS cate1
            , j.PT_Name AS cate2 
            FROM (SELECT PI_ID FROM (
            SELECT DISTINCT a.PI_ID FROM pic_distribute a
            JOIN user_distribute b ON a.E_User=b.E_User
            JOIN pic_index c ON a.PI_ID=c.PI_ID AND c.PI_Type=1
            WHERE c.PI_State=1 
            AND a.A_State =0 AND b."""+storage+"""=1
            AND a."""+storage+"""=0
            ORDER BY a.E_Time
            ) pic_distribute) e 
            LEFT JOIN pic_index d ON d.PI_ID=e.PI_ID
            INNER JOIN pic_priminfo f ON f.PI_ID=e.PI_ID  
            INNER JOIN pic_priminfo_en g ON g.PI_ID=e.PI_ID
            LEFT JOIN user_release h ON h.UR_ID=d.E_User
            INNER JOIN pic_type AS i ON g.PPI_Type1=i.PT_ID
            LEFT JOIN pic_type AS j ON g.PPI_Type2=j.PT_ID"""
            cursor.execute(sql)
            return cursor.fetchall()
    except Exception:
        print ('mysql error')

# 下载所有国际分发所需要的图片
def select_wait_distribute(connection):
    try:
        with connection.cursor() as cursor:
         # Read a single record
         sql = """SELECT DISTINCT(a.PI_ID),c.PI_SrcUrl,c.PI_FieldID,a.E_Time FROM pic_distribute a
         JOIN user_distribute b ON a.E_User=b.E_User
         JOIN pic_index c ON a.PI_ID=c.PI_ID AND c.PI_Type=1
         WHERE c.PI_State=1  
         AND a.A_State =0 
         ORDER BY a.E_Time"""
         cursor.execute(sql)
         results = cursor.fetchall()
         return results
    except Exception:
        print ('mysql error: 获取待分发图片列表失败')
dict = {}
connection = pymysql.connect(host='localhost', user='root', password='123456', db='origino', charset='utf8',cursorclass=pymysql.cursors.DictCursor)
pics = select_pics(connection, 'Pond5')
# pics = select_wait_distribute(connection)
connection.close()
for pic in pics:
    dict[pic['PI_ID']] =pic
wb = load_workbook('f:/pond5.xlsx')
ws = wb.active
for i in range(2,12025):
    pi_id = ws['B'+str(i)].value[:-4]
    p = dict[int(pi_id)]
    ws['E2'] = p['EN_Name']
    ws['K2'] = p['EN_Intro']
    print '修改第'+i+'行成功'
wb.save('new_pond5.xlsx')



